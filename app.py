import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
import pandas as pd
import os
from datetime import datetime

csv_file_path = '/Users/ibrahim/Desktop/My Work/ofis_otomasyon/persons.csv'

def format_time(time_str):
    try:
        for fmt in ('%H:%M', '%H.%M', '%H,%M'):
            try:
                return datetime.strptime(time_str, fmt).strftime('%H:%M')
            except ValueError:
                continue
        return time_str
    except ValueError:
        return time_str

def load_csv_data():
    try:
        df = pd.read_csv(csv_file_path)
        return df
    except Exception as e:
        messagebox.showerror("Hata", f"Dosya okunurken bir hata oluştu: {e}")
        return pd.DataFrame()

def autocomplete_name(event=None):
    query = entry_name.get().strip().lower()
    if not query:
        return

    df = load_csv_data()
    if df.empty:
        return

    results = df[df['persons'].str.lower().str.startswith(query, na=False)]
    if not results.empty:
        suggestions = results['persons'].tolist()
        if suggestions:
            listbox_suggestions.delete(0, tk.END)
            for suggestion in suggestions:
                listbox_suggestions.insert(tk.END, suggestion)
            listbox_suggestions.place(x=entry_name.winfo_x(), y=entry_name.winfo_y() + entry_name.winfo_height())
        else:
            listbox_suggestions.place_forget()
    else:
        listbox_suggestions.place_forget()

def set_suggestion(event=None):
    selected_suggestion = listbox_suggestions.get(tk.ACTIVE)
    entry_name.delete(0, tk.END)
    entry_name.insert(0, selected_suggestion)
    listbox_suggestions.place_forget()

def kaydet():
    full_name = entry_name.get()
    giris = entry_giris.get()
    cikis = entry_cikis.get()
    departman = combobox_departman.get()
    durum = combobox_durum.get()

    if not full_name or not giris or not cikis or not departman or departman == "Departman Seçiniz" or durum == "Durum Seçiniz":
        messagebox.showwarning("Uyarı", "Lütfen tüm alanları doldurduğunuzdan emin olun.")
        return

    full_name = ' '.join(word.capitalize() for word in full_name.split())

    today_date = datetime.now().strftime("%Y-%m-%d")
    giris = format_time(giris)
    cikis = format_time(cikis)

    giris_datetime = f"{today_date} {giris}"
    cikis_datetime = f"{today_date} {cikis}"

    file_name = f"/Users/ibrahim/Desktop/My Work/ofis_otomasyon/data/{today_date}_personel.xlsx"

    try:
        if not os.path.exists(file_name):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Personel Verisi"
            sheet.append(["Ad Soyad", "Giriş Tarih-Saat", "Çıkış Tarih-Saat", "Departman", "Durum"])
            workbook.save(file_name)

        workbook = load_workbook(file_name)
        sheet = workbook.active

        if durum == "Gelmedi":
            giris_datetime = cikis_datetime = "G"

        sheet.append([full_name, giris_datetime, cikis_datetime, departman, durum])
        workbook.save(file_name)

        messagebox.showinfo("Başarılı", "Veri başarıyla kaydedildi!")
        entry_name.delete(0, tk.END)
        entry_giris.delete(0, tk.END)
        entry_cikis.delete(0, tk.END)
        combobox_departman.set('Departman Seçiniz')
        combobox_durum.set('Durum Seçiniz')
        button_kaydet.config(bg="#CACFA4", fg="green")
    except Exception as e:
        messagebox.showerror("Hata", f"Bir hata oluştu: {e}")

# GUI oluşturma
app = tk.Tk()
app.title("Çalışan Yönetim Sistemi")
app.geometry("600x600")  # Pencere boyutunu büyütüyoruz
app.configure(bg="#f4f4f4")  # Arka plan rengini CSS'teki gibi ayarlıyoruz

main_frame = tk.Frame(app, padx=20, pady=20, bg="#fff")
main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

main_frame.columnconfigure(0, weight=1)
main_frame.rowconfigure(0, weight=1)

tk.Label(main_frame, text="Ad Soyad:", font=("Arial", 16), bg="#fff").grid(row=0, column=0, padx=15, pady=15, sticky=tk.W)
entry_name = tk.Entry(main_frame, width=40, font=("Arial", 16))
entry_name.grid(row=0, column=1, padx=15, pady=15, sticky=tk.E)
entry_name.bind("<KeyRelease>", autocomplete_name)

# Öneri kutusu için Listbox
listbox_suggestions = tk.Listbox(main_frame, height=5, width=40, font=("Arial", 16))
listbox_suggestions.place_forget()
listbox_suggestions.bind("<ButtonRelease-1>", set_suggestion)

tk.Label(main_frame, text="Giriş Saati (HH:MM):", font=("Arial", 16), bg="#fff").grid(row=1, column=0, padx=15, pady=15, sticky=tk.W)
entry_giris = tk.Entry(main_frame, width=40, font=("Arial", 16))
entry_giris.grid(row=1, column=1, padx=15, pady=15, sticky=tk.E)

tk.Label(main_frame, text="Çıkış Saati (HH:MM):", font=("Arial", 16), bg="#fff").grid(row=2, column=0, padx=15, pady=15, sticky=tk.W)
entry_cikis = tk.Entry(main_frame, width=40, font=("Arial", 16))
entry_cikis.grid(row=2, column=1, padx=15, pady=15, sticky=tk.E)

tk.Label(main_frame, text="Departman:", font=("Arial", 16), bg="#fff").grid(row=3, column=0, padx=15, pady=15, sticky=tk.W)
departmanlar = ["ML", "WEB", "metavers", "hukuk Stajyeri", "Stajyer (genel)"]
combobox_departman = ttk.Combobox(main_frame, values=departmanlar, width=38, state="readonly", font=("Arial", 16))
combobox_departman.set("Departman Seçiniz")
combobox_departman.grid(row=4, column=1, padx=15, pady=15, sticky=tk.E)

tk.Label(main_frame, text="Durum:", font=("Arial", 16), bg="#fff").grid(row=5, column=0, padx=15, pady=15, sticky=tk.W)
durumlar = ["Durum Seçiniz", "Geldi", "Gelmedi"]
combobox_durum = ttk.Combobox(main_frame, values=durumlar, width=38, state="readonly", font=("Arial", 16))
combobox_durum.set("Durum Seçiniz")
combobox_durum.grid(row=6, column=1, padx=15, pady=15, sticky=tk.E)

button_kaydet = tk.Button(main_frame, text="Kaydet", command=kaydet, font=("Arial", 16), bg="#28a745", fg="#fff")
button_kaydet.grid(row=7, column=0, columnspan=2, pady=20)

app.mainloop()
